#!/usr/bin/env python3
"""
Daily Cost Tracker
Logs estimated API and infrastructure costs to Excel
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path
from datetime import datetime, timedelta
import json
import os
import sys

EXCEL_PATH = Path(__file__).parent.parent / 'data' / 'cost-tracker.xlsx'
EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)

# Cost estimates (adjust as needed)
COSTS = {
    "aws_ec2_daily": 1.50,  # t3.small ~$45/month
    "opus_per_1k_input": 0.015,
    "opus_per_1k_output": 0.075,
    "sonnet_per_1k_input": 0.003,
    "sonnet_per_1k_output": 0.015,
}

def create_workbook():
    """Create new workbook with headers"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Costs"
    
    # Headers
    headers = ["Date", "Model Fees ($)", "AWS ($)", "Other ($)", "Total ($)", "Notes"]
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 40
    
    return wb

def load_or_create_workbook():
    """Load existing or create new workbook"""
    if EXCEL_PATH.exists():
        return openpyxl.load_workbook(EXCEL_PATH)
    return create_workbook()

def estimate_model_costs(date_str: str = None) -> tuple:
    """
    Estimate model costs for a date
    Returns (cost, notes)
    
    In future: parse actual usage from logs
    For now: estimate based on session patterns
    """
    # Check if we have session data
    # For now, use heuristics based on day patterns
    today = datetime.now()
    
    # Heavy session estimate (weekday with active work)
    # Light session estimate (minimal activity)
    # Default to medium estimate
    
    # TODO: Integrate with actual API usage tracking
    base_cost = 3.50  # Base daily (crons, heartbeats)
    
    notes = []
    
    # Check for interactive sessions (rough estimate)
    # Could parse memory files for activity indicators
    memory_dir = Path(__file__).parent.parent / 'memory'
    
    if date_str:
        date_file = memory_dir / f"{date_str}.md"
        if date_file.exists():
            content = date_file.read_text()
            lines = len(content.split('\n'))
            
            if lines > 200:
                base_cost += 15.0
                notes.append("Heavy session")
            elif lines > 100:
                base_cost += 8.0
                notes.append("Medium session")
            elif lines > 50:
                base_cost += 3.0
                notes.append("Light session")
    
    return round(base_cost, 2), ", ".join(notes) if notes else "Crons + heartbeats"

def add_entry(date_str: str, model_fees: float = None, aws: float = None, 
              other: float = 0, notes: str = ""):
    """Add a cost entry for a date"""
    wb = load_or_create_workbook()
    ws = wb.active
    
    # Find next row
    next_row = ws.max_row + 1
    
    # Check if date already exists
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == date_str:
            print(f"Entry for {date_str} already exists (row {row})")
            return False
    
    # Estimate costs if not provided
    if model_fees is None:
        model_fees, auto_notes = estimate_model_costs(date_str)
        if not notes:
            notes = auto_notes
    
    if aws is None:
        aws = COSTS["aws_ec2_daily"]
    
    total = model_fees + aws + other
    
    # Add row
    ws.cell(row=next_row, column=1, value=date_str)
    ws.cell(row=next_row, column=2, value=model_fees)
    ws.cell(row=next_row, column=3, value=aws)
    ws.cell(row=next_row, column=4, value=other)
    ws.cell(row=next_row, column=5, value=round(total, 2))
    ws.cell(row=next_row, column=6, value=notes)
    
    # Format numbers
    for col in [2, 3, 4, 5]:
        ws.cell(row=next_row, column=col).number_format = '#,##0.00'
    
    wb.save(EXCEL_PATH)
    print(f"Added entry: {date_str} | Models: ${model_fees:.2f} | AWS: ${aws:.2f} | Total: ${total:.2f}")
    return True

def show_summary():
    """Show cost summary"""
    if not EXCEL_PATH.exists():
        print("No cost data yet")
        return
    
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    
    total_model = 0
    total_aws = 0
    total_other = 0
    count = 0
    
    print("\n" + "=" * 60)
    print("COST TRACKER SUMMARY")
    print("=" * 60)
    
    for row in range(2, ws.max_row + 1):
        date = ws.cell(row=row, column=1).value
        model = ws.cell(row=row, column=2).value or 0
        aws = ws.cell(row=row, column=3).value or 0
        other = ws.cell(row=row, column=4).value or 0
        total = ws.cell(row=row, column=5).value or 0
        notes = ws.cell(row=row, column=6).value or ""
        
        print(f"{date}: ${total:.2f} ({notes[:30]})")
        
        total_model += model
        total_aws += aws
        total_other += other
        count += 1
    
    print("-" * 60)
    print(f"Total ({count} days): ${total_model + total_aws + total_other:.2f}")
    print(f"  Models: ${total_model:.2f}")
    print(f"  AWS:    ${total_aws:.2f}")
    print(f"  Other:  ${total_other:.2f}")
    print(f"Daily avg: ${(total_model + total_aws + total_other) / max(count, 1):.2f}")

def main():
    if len(sys.argv) < 2:
        print("Usage:")
        print("  cost-tracker.py add <date> [model_fees] [aws] [other] [notes]")
        print("  cost-tracker.py yesterday       # Add yesterday's entry")
        print("  cost-tracker.py today           # Add today's entry")
        print("  cost-tracker.py summary         # Show summary")
        print("  cost-tracker.py path            # Show Excel path")
        sys.exit(1)
    
    cmd = sys.argv[1].lower()
    
    if cmd == "yesterday":
        yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        add_entry(yesterday)
    
    elif cmd == "today":
        today = datetime.now().strftime("%Y-%m-%d")
        add_entry(today)
    
    elif cmd == "summary":
        show_summary()
    
    elif cmd == "path":
        print(EXCEL_PATH)
    
    elif cmd == "add":
        if len(sys.argv) < 3:
            print("Need date: cost-tracker.py add 2026-02-02")
            sys.exit(1)
        
        date_str = sys.argv[2]
        model_fees = float(sys.argv[3]) if len(sys.argv) > 3 else None
        aws = float(sys.argv[4]) if len(sys.argv) > 4 else None
        other = float(sys.argv[5]) if len(sys.argv) > 5 else 0
        notes = sys.argv[6] if len(sys.argv) > 6 else ""
        
        add_entry(date_str, model_fees, aws, other, notes)
    
    else:
        print(f"Unknown command: {cmd}")
        sys.exit(1)

if __name__ == "__main__":
    main()
